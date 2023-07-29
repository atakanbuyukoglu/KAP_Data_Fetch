from .RequestWrapper import Request
from .DataQuery import YahooSession
import json
from pathlib import Path
from bs4 import BeautifulSoup
import re
import numpy as np
import time
import pandas as pd
from yahooquery import Ticker
import pickle as pk
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties


COMPANIES_FILE = "Companies.pickle"
COMPANIES_PATH = Path(__file__).parents[1] / 'Data' / 'Companies'
COMPANIES_PATH.mkdir(parents=True, exist_ok=True)

# TODO: Store All Values in CSV files instead of pickle
class KAP():

    def __init__(self) -> None:
        self.constants_path = Path(__file__).parents[1] / 'Config' / 'Constants.json'
        # Load the constants
        with open(self.constants_path, 'r') as f:
            self.constants = json.load(f)
        self.data_path = Path(__file__).parents[1] / 'Data'
        self.companies = None

        # Yahoo Finance variables
        self.query_suffix = '.IS'
        self.yahoo_session = YahooSession(sleep_time=self.constants['sleep_times']['yahoo'])

        # Initialize Request wrapper with sleep time
        self.r = Request(sleep_time=self.constants['sleep_times']['kap'])

    def __get_company_list_html(self):

        # Load the constants
        with open(self.constants_path, 'r') as f:
            constants = json.load(f)
        
        # Get the site
        company_list_site = constants['sites']['company_list']
        
        # Fetch the data from the site
        raw_response = self.r.get(company_list_site)
        # Check the status code
        raw_response.raise_for_status()

        # String for compatibility with offline version
        raw_response = raw_response.text

        # Save the HTML results to a log file
        with open(self.data_path / 'Company_List.html', 'w', encoding='utf-8') as f:
            f.write(raw_response)
    
        return raw_response

    def __raw_html_to_html_list(self, raw_html):
        soup = BeautifulSoup(raw_html, 'html.parser')
        html_list = soup.find_all('div', 'w-clearfix w-inline-block comp-row')

        return html_list

    def __html_list_to_company_dict(self, html_list, old_company_dict=None, save_results=True):
        # Load the constants
        with open(self.constants_path, 'r') as f:
            constants = json.load(f)

        company_dict = {}
        for company in html_list:
            ticker_dict = {}

            ticker = company.select('div.comp-cell._04.vtable a.vcell')[0].text
            ticker_dict['ticker'] = ticker
            name = company.select('div.comp-cell._14.vtable a.vcell')[0].text
            ticker_dict['name'] = name
            link = company.select('div.comp-cell._04.vtable a.vcell[href]')[0]['href']
            ticker_dict['link'] = constants['sites']['main_site'] + link
            auditor = company.select('div.comp-cell._11.vtable a.vcell')[0].text
            ticker_dict['auditor'] = auditor
            city = company.select('div.comp-cell._12.vtable div.vcell')[0].text
            ticker_dict['city'] = city
            kap_id = int(re.search(r'\d+', link).group())
            ticker_dict['kap_id'] = kap_id

            # Add values from old list if there are any#
            if old_company_dict is not None and ticker in old_company_dict:
                for old_key, old_value in old_company_dict[ticker].items():
                    if old_key not in ticker_dict:
                        ticker_dict[old_key] = old_value

            company_dict[ticker] = ticker_dict
        
        if save_results:
            self.save_companies(company_dict)

        return company_dict

    def save_companies(self, companies):
        with open(self.data_path / COMPANIES_FILE, 'wb') as f:
            pk.dump(companies, f)
    
    def save_company(self, company_dict):
        companies = self.get_companies()
        companies[company_dict['ticker']] = company_dict
        self.save_companies(companies)
        # Save to the Excel files
        wb = Workbook()
        ws_info = wb.active
        ws_info.title = 'Info'
        ws_info.sheet_properties.best_fit = True
        company_ticker = company_dict['ticker']
        for company_key, company_value in company_dict.items():
            if isinstance(company_value, pd.DataFrame):
                ws_df = wb.create_sheet(company_key.upper())
                for row in dataframe_to_rows(company_value, index=True, header=True):
                    ws_df.append(row)
            elif isinstance(company_value, dict):
                stats = company_value[(company_ticker + self.query_suffix)]
                ws_stats = wb.create_sheet(company_key.upper())
                for stat, value in stats.items():
                    ws_stats.append([stat, value])
            else:
                ws_info.append([company_key, company_value])

        wb.save(COMPANIES_PATH / (company_ticker + '.xlsx'))

    # TODO: Load company data from XLSX files
    def load_company(self, ticker, online=False):
        ticker = ticker.upper()
        company_info = {}
        # Load the company info from the xlsx file
        company_path = COMPANIES_PATH / (ticker + '.xlsx')
        wb = load_workbook(company_path)
        # The Info worksheet
        ws_info = wb.get_sheet_by_name('Info')
        for r in ws_info.rows:
            company_info[r[0].value] = r[1].value
        # The stats worksheet
        ws_stats = wb.get_sheet_by_name('KEY_STATS')
        stats_ticker = ticker + self.query_suffix
        stats_ticker_dict = {}
        for r in ws_stats.rows:
            stats_ticker_dict[r[0].value] = r[1].value
        stats_dict = {stats_ticker : stats_ticker_dict}
        company_info['key_stats'] = stats_dict
        # The financial information worksheets
        raise NotImplementedError

        return company_info

    def update_company(self, ticker, online=True):
        # Get the company dictionary
        company_info = self.get_company(ticker)
        mkk_id = self.get_mkk_id(ticker)
        if online:
            # Load the report indices
            company_filter_site = self.constants['sites']['filter_site']
            company_filter_site += '/' + mkk_id + '/FR/365'
            reports_text = self.r.get(company_filter_site)
            reports_json = json.loads(reports_text.text)
            # Filter the financial reports
            reports_financial = [report for report in reports_json if report['basic']['disclosureCategory'] == 'FR']
            report_indices = [report['basic']['disclosureIndex'] for report in reports_financial]
            company_info['report_indices'] = report_indices

        return company_info

    
    def update_companies(self, ticker, online=True):
        if online:
            # Update all financial information regarding the ticker
            self.get_balance_sheet(ticker)
            self.get_income_statement(ticker)
            self.get_cash_flow(ticker)
        # Return an updated list of companies
        return self.get_companies(online=online)

    def get_mkk_id(self, ticker:str):
        company_info = self.get_company(ticker)
        # Try returning the MKK ID if it already exists
        try:
            return company_info['mkk_id']
        # Get the MKK ID from the KAP website instead
        except KeyError:            
            resp = self.r.get(url=company_info['link'])
            soup = BeautifulSoup(resp.text, 'html.parser')
            mkk_id = soup.select('img.comp-logo')[0]['src'].split('/')[-1]
            company_info['mkk_id'] = mkk_id
            self.save_company(company_info)
            return mkk_id

        query_ticker = ticker + self.query_suffix
        ticker_obj = Ticker(query_ticker)
        return ticker_obj.summary_detail

    def get_yahoo_property(self, ticker:str, property:str, online:bool=False):
        # Get the company info
        company_info = self.get_company(ticker)
        # Force online update here
        if online:
            query_ticker = ticker + self.query_suffix
            query = Ticker(query_ticker)
            # Exception is raised here if the property does not exist
            company_info[property] = getattr(query, property)
            self.save_company(company_info)
            return company_info[property]
        else:
            # Try returning the stats if it already exists
            try:
                return company_info[property]
            # Get the stats from the Yahoo instead
            except KeyError:
                return self.get_yahoo_property(ticker, property, online=True)

    def get_price(self, ticker:str):
        query_ticker = ticker + self.query_suffix
        financial_data = self.get_yahoo_property(ticker, 'financial_data', online=True)
        return financial_data[query_ticker]['currentPrice']

    # TODO: Get these info from KAP
    def get_balance_sheet(self, ticker:str, frequency='q', trailing=True, online:bool=False):
        # Get the company info
        company_info = self.get_company(ticker)
        # Force online update here
        if online:
            query_ticker = ticker + self.query_suffix
            query = Ticker(query_ticker)
            company_info['balance_sheet'] = query.balance_sheet(frequency=frequency, trailing=trailing)
            self.save_company(company_info)
            return company_info['balance_sheet']
        else:
            # Try returning the balance_sheet if it already exists
            try:
                return company_info['balance_sheet']
            # Get the balance_sheet from the Yahoo instead
            except KeyError:
                return self.get_balance_sheet(ticker, frequency=frequency, trailing=trailing, online=True)

    def get_cash_flow(self, ticker:str, frequency='q', trailing=True, online:bool=False):
        # Get the company info
        company_info = self.get_company(ticker)
        # Force online update here
        if online:
            query_ticker = ticker + self.query_suffix
            query = Ticker(query_ticker)
            company_info['cash_flow'] = query.cash_flow(frequency=frequency, trailing=trailing)
            self.save_company(company_info)
            return company_info['cash_flow']
        else:
            # Try returning the cash_flow if it already exists
            try:
                return company_info['cash_flow']
            # Get the cash_flow from the Yahoo instead
            except KeyError:
                return self.get_cash_flow(ticker, frequency=frequency, trailing=trailing, online=True)
            
    def get_income_statement(self, ticker:str, frequency='q', trailing=True, online:bool=False):
        # Get the company info
        company_info = self.get_company(ticker)
        # Force online update here
        if online:
            query_ticker = ticker + self.query_suffix
            query = Ticker(query_ticker)
            company_info['income_statement'] = query.income_statement(frequency=frequency, trailing=trailing)
            self.save_company(company_info)
            return company_info['income_statement']
        else:
            # Try returning the income_statement if it already exists
            try:
                return company_info['income_statement']
            # Get the income_statement from the Yahoo instead
            except KeyError:
                return self.get_income_statement(ticker, frequency=frequency, trailing=trailing, online=True)

    def get_balance_sheet_kap(self, ticker:str, frequency='q', trailing=True, online:bool=False):
        mkk_id = self.get_mkk_id(ticker)
        if online:
            # Load the report indices
            company_filter_site = self.constants['sites']['filter_site']
            company_filter_site += '/' + mkk_id + '/FR/365'
            reports_text = self.r.get(company_filter_site)
            reports_json = json.loads(reports_text.text)
            # Filter the financial reports
            reports_financial = [report for report in reports_json if report['basic']['disclosureCategory'] == 'FR']
            report_indices = [report['basic']['disclosureIndex'] for report in reports_financial]
            

    def get_company(self, ticker:str, online:bool=False):
        companies = self.get_companies(online=online)
        if ticker not in companies:
            companies = self.get_companies(online=True)
        # Try accessing the ticker
        try:
            company_info = companies[ticker]
            return company_info
        except KeyError as e:
            print("Ticker", ticker, "could not be found in the companies list.")
            raise e

    def get_companies(self, online=False, save_results=True):

        if online:
            # Get the old version of companies
            # Get it from the saved varible if possible
            if self.companies is not None:
                companies_old = self.companies
            # If not there, try loading them from the saved file
            else:
                try:
                    with open(self.data_path / COMPANIES_FILE, 'rb') as f:
                        companies_old = pk.load(f)
                except FileNotFoundError:
                    companies_old = None
            
            # Get the raw HTML response from KAP
            raw_response = self.__get_company_list_html()

            # Filter the results to companies
            html_list = self.__raw_html_to_html_list(raw_response)
            
            # Convert the results to a dictionary
            companies = self.__html_list_to_company_dict(html_list, old_company_dict=companies_old,
                                                            save_results=save_results)
        else:
            # Get it from the saved variable if possible
            if self.companies is not None:
                return self.companies
            # Try loading them from the saved file
            try:
                with open(self.data_path / COMPANIES_FILE, 'rb') as f:
                    companies = pk.load(f)
            # If the file is not there, obtain it online
            except FileNotFoundError:
                return self.get_companies(online=True, save_results=save_results)
        
        self.companies = companies
        return companies

    def get_query_ticker(self, ticker):
        return ticker + self.query_suffix

